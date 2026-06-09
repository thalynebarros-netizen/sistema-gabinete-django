from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from contacts.importing import (
    HEADER_ALIASES,
    build_external_id,
    bulk_save_contacts,
    clean_phone,
    clean_value,
    contact_from_indexes,
    infer_state,
    load_contact_indexes,
    normalize_label,
    pick_value,
    sync_contact_fields,
)
from contacts.models import Contact

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - depends on local environment.
    raise CommandError(
        "A biblioteca openpyxl nao esta instalada. Rode pip install -r requirements.txt."
    ) from exc


CONTACT_SHEET_NAME = "Base Completa"


def find_sheet(workbook, requested_name):
    requested = normalize_label(requested_name)
    for worksheet in workbook.worksheets:
        sheet_name = normalize_label(worksheet.title)
        if sheet_name == requested or requested in sheet_name:
            return worksheet
    available = ", ".join(workbook.sheetnames)
    raise CommandError(
        f"Nao encontrei a aba '{requested_name}'. Abas disponiveis: {available}"
    )


def find_header_row(worksheet):
    expected_names = {
        normalize_label(alias)
        for field in ("external_id", "full_name", "phone")
        for alias in HEADER_ALIASES[field]
    }
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=15, values_only=True),
        start=1,
    ):
        normalized = {normalize_label(value) for value in row if clean_value(value)}
        if normalize_label("id") in normalized and normalized.intersection(expected_names):
            return row_number
    raise CommandError(
        "Nao encontrei o cabecalho da aba de contatos. Verifique se ha colunas ID, nome e telefone."
    )


def build_payload(headers, row):
    payload = {}
    for index, header in enumerate(headers):
        key = clean_value(header) or f"coluna_{index + 1}"
        payload[key] = clean_value(row[index]) if index < len(row) else ""
    return payload


class Command(BaseCommand):
    help = "Importa contatos da aba Base Completa de uma planilha Excel."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", help="Caminho da planilha Excel principal.")
        parser.add_argument(
            "--sheet",
            default=CONTACT_SHEET_NAME,
            help="Nome da aba com os contatos. O padrao procura por 'Base Completa'.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Simula a importacao e desfaz as alteracoes no final.",
        )

    def handle(self, *args, **options):
        workbook_path = Path(options["xlsx_path"]).expanduser()
        if not workbook_path.exists():
            raise CommandError(f"Planilha nao encontrada: {workbook_path}")

        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        worksheet = find_sheet(workbook, options["sheet"])
        header_row = find_header_row(worksheet)
        headers = next(
            worksheet.iter_rows(
                min_row=header_row,
                max_row=header_row,
                values_only=True,
            )
        )

        totals = {"created": 0, "updated": 0, "skipped": 0}
        with transaction.atomic():
            contacts_by_external, contacts_by_phone = load_contact_indexes()
            created_contacts = []
            updated_contacts = {}
            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=header_row + 1, values_only=True),
                start=header_row + 1,
            ):
                if not any(clean_value(value) for value in row):
                    continue

                payload = build_payload(headers, row)
                full_name = pick_value(payload, "full_name")
                phone = clean_phone(pick_value(payload, "phone"))
                if not full_name:
                    totals["skipped"] += 1
                    continue

                external_id = build_external_id(payload, row_number)
                fields = {
                    "external_id": external_id,
                    "full_name": full_name,
                    "phone": phone,
                    "city": pick_value(payload, "city"),
                    "state": infer_state(payload),
                    "neighborhood": pick_value(payload, "neighborhood"),
                    "profile": pick_value(payload, "profile"),
                    "source": pick_value(payload, "source"),
                    "notes_summary": pick_value(payload, "notes_summary"),
                    "source_payload": {
                        "workbook": workbook_path.name,
                        "sheet": worksheet.title,
                        "row_number": row_number,
                        "responsavel": payload.get("RESPONSAVEL", "")
                        or payload.get("RESPONSÁVEL", ""),
                        "row": payload,
                    },
                }

                contact = contact_from_indexes(
                    external_id,
                    phone,
                    contacts_by_external,
                    contacts_by_phone,
                )
                if contact is None:
                    contact = Contact(**fields)
                    created_contacts.append(contact)
                    contacts_by_external[external_id] = contact
                    if phone:
                        contacts_by_phone[phone] = contact
                    totals["created"] += 1
                    continue

                if sync_contact_fields(contact, fields):
                    if contact.pk:
                        updated_contacts[contact.pk] = contact
                    totals["updated"] += 1

            bulk_save_contacts(created_contacts, updated_contacts)
            if options["dry_run"]:
                transaction.set_rollback(True)

        mode = "Simulacao" if options["dry_run"] else "Importacao"
        self.stdout.write(
            self.style.SUCCESS(
                f"{mode} concluida: {totals['created']} criados, "
                f"{totals['updated']} atualizados, {totals['skipped']} ignorados."
            )
        )
        self.stdout.write(
            f"Aba usada: {options['sheet']}. Cabecalho na linha {header_row}."
        )
